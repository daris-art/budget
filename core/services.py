# core/services.py

import json
import openpyxl
import requests
import datetime
import logging
from pathlib import Path
from typing import List
from dataclasses import asdict
from openpyxl.worksheet.worksheet import Worksheet
from core.database import DatabaseManager
from core.data_models import Result, Depense, DatabaseError

logger = logging.getLogger(__name__)

class BitcoinAPIService:
    """Service pour récupérer le prix du Bitcoin."""
    def get_price(self) -> Result:
        url = "https://api.coingecko.com/api/v3/simple/price"
        params = {"ids": "bitcoin", "vs_currencies": "eur"}
        try:
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            price = data.get("bitcoin", {}).get("eur")
            
            if price is None:
                return Result.error("Format de réponse de l'API inattendu.")
            return Result.success(data=price)
        except requests.exceptions.RequestException as e:
            logger.error(f"Erreur réseau BTC: {e}")
            return Result.error("Erreur réseau. Vérifiez votre connexion.")
        except Exception as e:
            logger.error(f"Erreur API BTC: {e}")
            return Result.error("Une erreur inattendue est survenue.")


class ImportExportService:
    """Service pour l'import/export de fichiers."""
    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager

    def export_to_json(self, mois_id: int, filepath: Path) -> Result:
        """Exporte les données d'un mois (salaire et dépenses) vers un fichier JSON."""
        try:
            mois_details = self.db_manager.get_mois_by_id(mois_id)
            if not mois_details:
                return Result.error("Mois non trouvé pour l'export.")
            
            depenses = self.db_manager.get_depenses_by_mois(mois_id)
            
            # --- CORRECTION ---
            # On utilise asdict(dep) pour les dataclasses au lieu de dep._asdict()
            depenses_data = [asdict(dep) for dep in depenses]
            
            # Structure des données pour le fichier JSON
            data_to_export = {
                "mois": {
                    "nom": mois_details.nom,
                    "salaire": mois_details.salaire
                },
                "depenses": depenses_data
            }

            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data_to_export, f, indent=4, ensure_ascii=False)
            
            return Result.success(f"Export réussi vers {filepath.name}")

        except Exception as e:
            logger.error(f"Erreur inattendue lors de l'export JSON: {e}")
            return Result.error(f"Une erreur inattendue est survenue: {e}")

    def import_from_json(self, filepath: Path, new_mois_name: str) -> Result:
        """Importe un fichier JSON pour créer un nouveau mois et ses dépenses."""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if 'depenses' not in data:
                return Result.error("Structure JSON invalide : clé 'depenses' manquante.")

            salaire = data.get('mois', {}).get('salaire', 0.0)
            
            depenses_a_importer = []
            for dep_data in data['depenses']:
                depenses_a_importer.append(Depense(
                    nom=dep_data.get('nom', 'Dépense sans nom'),
                    montant=dep_data.get('montant', 0.0),
                    categorie=dep_data.get('categorie', 'Autres'),
                    date_depense=dep_data.get('date_depense', datetime.datetime.now().strftime('%d/%m/%Y')),
                    est_credit=dep_data.get('est_credit', False),
                    effectue=dep_data.get('effectue', False),
                    emprunte=dep_data.get('emprunte', False)
                ))
            
            # On appelle la méthode transactionnelle
            self.db_manager.import_new_mois(new_mois_name, salaire, depenses_a_importer)
            
            return Result.success(f"Import réussi: {len(depenses_a_importer)} dépenses ajoutées à '{new_mois_name}'.")

        except FileNotFoundError:
            return Result.error("Fichier non trouvé.")
        except json.JSONDecodeError as e:
            return Result.error(f"Erreur de décodage JSON : {e}")
        except DatabaseError as e:
            return Result.error(str(e)) # L'erreur de la DB remonte ici
        except Exception as e:
            logger.error(f"Erreur inattendue lors de l'import JSON: {e}")
            return Result.error(f"Une erreur inattendue est survenue: {e}")

    def import_from_excel(self, filepath: Path, new_mois_name: str, progress_callback=None) -> Result:
        """
        Lit un fichier Excel, crée un nouveau mois et y importe les opérations.
        Les crédits sont sommés et définissent le salaire initial du mois.
        """
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            sheet: Worksheet = workbook.active
            
            header_row_index = -1
            col_indices = {}

            # Recherche de la ligne d'en-tête à partir de la ligne 10
            for i in range(1, sheet.max_row + 1): # On commence à la ligne 1 pour plus de flexibilité
                row_values = [str(cell.value).strip() if cell.value else "" for cell in sheet[i]]
                if "Libellé" in row_values and "Date" in row_values:
                    header_row_index = i
                    col_indices["nom"] = row_values.index("Libellé")
                    col_indices["date"] = row_values.index("Date")
                    if "Débit euros" in row_values:
                        col_indices["debit"] = row_values.index("Débit euros")
                    if "Crédit euros" in row_values:
                        col_indices["credit"] = row_values.index("Crédit euros")
                    break
            
            if header_row_index == -1:
                return Result.error("En-tête non trouvé. Vérifiez que les colonnes 'Libellé' et 'Date' existent.")

            # On lit d'abord toutes les lignes en mémoire pour connaître le total
            rows_to_process = list(sheet.iter_rows(min_row=header_row_index + 1))
            total_rows = len(rows_to_process)
            if total_rows == 0:
                return Result.error("Aucune donnée trouvée après l'en-tête.")

            operations_a_importer: List[Depense] = []
            
            # On parcourt les lignes pré-chargées
            for row in rows_to_process: 

                cells = [cell.value for cell in row]
                nom = cells[col_indices["nom"]]
                date_val = cells[col_indices["date"]]
                debit_val = cells[col_indices.get("debit", -1)] if "debit" in col_indices else None
                credit_val = cells[col_indices.get("credit", -1)] if "credit" in col_indices else None

                if not nom or not date_val:
                    continue

                date_depense_str = ""
                if isinstance(date_val, datetime.datetime):
                    date_depense_str = date_val.strftime('%d/%m/%Y')
                else:
                    date_depense_str = str(date_val)

                montant, est_credit = 0.0, False
                try:
                    if "credit" in col_indices and credit_val and float(credit_val) > 0:
                        montant, est_credit = float(credit_val), True
                    elif "debit" in col_indices and debit_val and float(debit_val) > 0:
                        montant, est_credit = float(debit_val), False
                    else:
                        continue
                except (ValueError, TypeError):
                    logger.warning(f"Ligne {i + header_row_index + 1} ignorée: montant invalide.")
                    continue
                
                operations_a_importer.append(
                    Depense(nom=str(nom).strip(), montant=montant, date_depense=date_depense_str, est_credit=est_credit, effectue= True)
                )
            
            # 1. On calcule le total des crédits qui servira de salaire initial
            salaire_initial = sum(op.montant for op in operations_a_importer if op.est_credit)

            # 2. On appelle notre méthode transactionnelle unique
            self.db_manager.import_new_mois(new_mois_name, salaire_initial, operations_a_importer)


            return Result.success(f"{len(operations_a_importer)} opérations importées dans '{new_mois_name}'.")

        except FileNotFoundError:
            return Result.error("Fichier Excel non trouvé.")
        except Exception as e:
            logger.error(f"Erreur inattendue lors de l'import Excel: {e}")
            return Result.error(f"Une erreur inattendue est survenue: {e}")
