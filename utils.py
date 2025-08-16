# utils.py - Classes de support pour l'architecture améliorée

import sqlite3
import json
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import List, Optional, Tuple, Any, Dict
import datetime
import logging
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from dataclasses import dataclass, asdict


# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ===== EXCEPTIONS PERSONNALISÉES =====
class BudgetAppError(Exception):
    """Exception de base de l'application"""
    pass

class DatabaseError(BudgetAppError):
    """Erreurs liées à la base de données"""
    pass

class ValidationError(BudgetAppError):
    """Erreurs de validation"""
    pass

class ImportExportError(BudgetAppError):
    """Erreurs d'import/export"""
    pass

# ===== CLASSES DE RÉSULTAT =====
@dataclass
class Result:
    """Classe pour encapsuler les résultats d'opérations"""
    is_success: bool
    message: str = ""
    error: str = ""
    data: Any = None
    
    @classmethod
    def success(cls, message: str = "", data: Any = None) -> 'Result':
        return cls(is_success=True, message=message, data=data)
    
    @classmethod
    def error(cls, error: str) -> 'Result':
        return cls(is_success=False, error=error)

@dataclass
class ValidationResult:
    """Résultat de validation des données"""
    is_valid: bool
    errors: List[str]
    validated_data: Dict[str, Any] = None
    
    def __post_init__(self):
        if self.validated_data is None:
            self.validated_data = {}

# ===== ENTITÉS DE DONNÉES =====
@dataclass
@dataclass
class Depense:
    """DTO pour une dépense ou une opération."""
    # On ajoute des valeurs par défaut pour plus de flexibilité
    id: Optional[int] = None
    nom: str = ''
    montant: float = 0.0
    categorie: str = 'Autres'
    date_depense: str = ''
    est_credit: bool = False
    
    effectue: bool = False
    emprunte: bool = False


@dataclass
class Mois:
    nom: str
    salaire: float = 0.0
    date_creation: Optional[str] = None
    id: Optional[int] = None

@dataclass
class MoisInput:
    """DTO pour l'input de création de mois"""
    nom: str
    salaire: str

@dataclass
class MoisDisplayData:
    """DTO pour l'affichage des données de mois"""
    nom: str
    salaire: float
    depenses: List[Depense]
    nombre_depenses: int  # <-- AJOUT DE CETTE LIGNE
    total_depenses: float
    argent_restant: float
    total_effectue: float
    total_non_effectue: float
    total_emprunte: float

# ===== VALIDATEUR DE DONNÉES =====
class DataValidator:
    """Validation centralisée des données"""
    
    @staticmethod
    def validate_mois_data(nom: str, salaire: str) -> ValidationResult:
        """Valide les données de création d'un mois"""
        errors = []
        validated_data = {}
        
        # Validation du nom
        if not nom or not nom.strip():
            errors.append("Le nom du mois est requis")
        else:
            validated_data['nom'] = nom.strip()
        
        # Validation du salaire
        try:
            validated_salaire = float(salaire.replace(',', '.')) if salaire else 0.0
            if validated_salaire < 0:
                errors.append("Le salaire ne peut pas être négatif")
            validated_data['salaire'] = validated_salaire
        except (ValueError, AttributeError):
            errors.append("Le salaire doit être un nombre valide")
            validated_data['salaire'] = 0.0
        
        return ValidationResult(
            is_valid=len(errors) == 0,
            errors=errors,
            validated_data=validated_data
        )
    
    @staticmethod
    def validate_expense_data(nom: str, montant: str, categorie: str) -> ValidationResult:
        """Valide les données d'une dépense"""
        errors = []
        validated_data = {}
        
        # Validation du nom (optionnel)
        validated_data['nom'] = nom.strip() if nom else ""
        
        # Validation du montant
        try:
            validated_montant = float(montant.replace(',', '.')) if montant else 0.0
            if validated_montant < 0:
                errors.append("Le montant ne peut pas être négatif")
            validated_data['montant'] = validated_montant
        except (ValueError, AttributeError):
            errors.append("Le montant doit être un nombre valide")
            validated_data['montant'] = 0.0
        
        # Validation de la catégorie
        validated_data['categorie'] = categorie if categorie else "Autres"
        
        return ValidationResult(
            is_valid=len(errors) == 0,
            errors=errors,
            validated_data=validated_data
        )

# ===== GESTIONNAIRE DE BASE DE DONNÉES =====
class DatabaseManager:
    """Responsable uniquement des opérations SQLite"""
    
    def __init__(self, db_path: Path):
        self.db_path = db_path
        self._init_database()
    
    def _init_database(self):
        """Initialise la base de données et crée les tables"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS mois (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        nom TEXT UNIQUE NOT NULL,
                        salaire REAL NOT NULL DEFAULT 0.0,
                        date_creation TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                
                # --- MODIFICATION DE LA TABLE 'depenses' ---
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS depenses (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        mois_id INTEGER NOT NULL,
                        nom TEXT NOT NULL DEFAULT '',
                        montant REAL NOT NULL DEFAULT 0.0,
                        categorie TEXT NOT NULL DEFAULT 'Autres',
                        
                        date_depense TEXT NOT NULL DEFAULT (strftime('%d/%m/%Y', 'now')),
                        
                        -- AJOUT : Champ pour indiquer si c'est un crédit ou un débit
                        est_credit BOOLEAN NOT NULL DEFAULT 0, -- 0 = débit (dépense), 1 = crédit (revenu)
                        
                        effectue BOOLEAN NOT NULL DEFAULT 0,
                        emprunte BOOLEAN NOT NULL DEFAULT 0,
                        FOREIGN KEY (mois_id) REFERENCES mois (id) ON DELETE CASCADE
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS config (
                        cle TEXT PRIMARY KEY,
                        valeur TEXT NOT NULL
                    )
                ''')
                
                conn.commit()
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de l'initialisation de la base de données: {e}")
    
    def create_mois(self, nom: str, salaire: float) -> int:
        """Crée un mois et retourne son ID"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    'INSERT INTO mois (nom, salaire) VALUES (?, ?)',
                    (nom, salaire)
                )
                mois_id = cursor.lastrowid
                conn.commit()
                return mois_id
        except sqlite3.IntegrityError:
            raise DatabaseError(f"Le mois '{nom}' existe déjà")
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la création du mois: {e}")
    
    def get_all_mois(self) -> List[Mois]:
        """Récupère tous les mois"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT id, nom, salaire, date_creation FROM mois ORDER BY date_creation DESC')
                rows = cursor.fetchall()
                return [Mois(nom=row[1], salaire=row[2], date_creation=row[3], id=row[0]) for row in rows]
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la récupération des mois: {e}")
    
    def get_mois_by_name(self, nom: str) -> Optional[Mois]:
        """Récupère un mois par son nom"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT id, nom, salaire, date_creation FROM mois WHERE nom = ?', (nom,))
                row = cursor.fetchone()
                if row:
                    return Mois(nom=row[1], salaire=row[2], date_creation=row[3], id=row[0])
                return None
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la récupération du mois: {e}")
    
    # Dans utils.py, à l'intérieur de la classe DatabaseManager, remplacez cette méthode :

    def delete_mois(self, nom: str) -> bool:
        """
        Supprime un mois et toutes ses dépenses associées en deux étapes explicites
        au sein d'une seule transaction.
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()

                # Étape 1 : Trouver l'ID du mois à partir de son nom.
                cursor.execute('SELECT id FROM mois WHERE nom = ?', (nom,))
                row = cursor.fetchone()

                # Si le mois n'est pas trouvé, on ne peut rien supprimer.
                if not row:
                    return False

                mois_id = row[0]

                # Étape 2 : Supprimer toutes les dépenses de la table 'depenses'
                # qui sont associées à cet ID de mois.
                cursor.execute('DELETE FROM depenses WHERE mois_id = ?', (mois_id,))
                
                # Étape 3 : Une fois les dépenses supprimées, supprimer le mois lui-même.
                cursor.execute('DELETE FROM mois WHERE id = ?', (mois_id,))

                # La transaction est validée automatiquement à la fin du bloc 'with'
                conn.commit()

                # On retourne True si la suppression du mois a bien eu lieu (rowcount > 0)
                return cursor.rowcount > 0

        except sqlite3.Error as e:
            # En cas d'erreur, la transaction est automatiquement annulée.
            raise DatabaseError(f"Erreur lors de la suppression du mois et de ses dépenses : {e}")
            
    def update_mois_salaire(self, mois_id: int, salaire: float):
        """Met à jour le salaire d'un mois"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('UPDATE mois SET salaire = ? WHERE id = ?', (salaire, mois_id))
                conn.commit()
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la mise à jour du salaire: {e}")
        
    def update_mois_name(self, mois_id: int, new_name: str):
        """Met à jour le nom d'un mois."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('UPDATE mois SET nom = ? WHERE id = ?', (new_name, mois_id))
                conn.commit()
        except sqlite3.IntegrityError:
            raise DatabaseError(f"Le nom '{new_name}' existe déjà.")
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors du renommage du mois: {e}")

    def get_depenses_by_mois(self, mois_id: int) -> List[Depense]:
        """Récupère toutes les dépenses pour un mois donné."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # --- CORRECTION ---
                # On ajoute 'date_depense' et 'est_credit' à la requête SELECT
                sql = '''
                    SELECT id, nom, montant, categorie, date_depense, est_credit, effectue, emprunte 
                    FROM depenses WHERE mois_id = ?
                '''
                
                cursor.execute(sql, (mois_id,))
                rows = cursor.fetchall()
                
                depenses = []
                for row in rows:
                    # On s'assure que les valeurs correspondent à l'ordre des colonnes
                    depenses.append(Depense(
                        id=row[0],
                        nom=row[1],
                        montant=row[2],
                        categorie=row[3],
                        date_depense=row[4],
                        est_credit=bool(row[5]), # Conversion en booléen
                        effectue=bool(row[6]),   # Conversion en booléen
                        emprunte=bool(row[7])    # Conversion en booléen
                    ))
                return depenses
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la récupération des dépenses: {e}")
        
    def create_depense(self, mois_id: int, depense: Depense) -> int:
        """Crée une nouvelle dépense dans la base de données et retourne son ID."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # --- CORRECTION ---
                # On ajoute les nouvelles colonnes à la requête SQL
                sql = '''
                    INSERT INTO depenses (
                        mois_id, nom, montant, categorie, 
                        date_depense, est_credit, effectue, emprunte
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                '''
                
                # On s'assure que les valeurs correspondent aux colonnes
                values = (
                    mois_id,
                    depense.nom,
                    depense.montant,
                    depense.categorie,
                    depense.date_depense,
                    depense.est_credit,
                    depense.effectue,
                    depense.emprunte
                )
                
                cursor.execute(sql, values)
                conn.commit()
                return cursor.lastrowid
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la création de la dépense : {e}")
 
    def update_depense(self, depense: Depense):
        """Met à jour une dépense existante dans la base de données."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                # --- MODIFICATION : Ajout de date_depense à la requête UPDATE ---
                sql = '''
                    UPDATE depenses SET
                        nom = ?, montant = ?, categorie = ?, date_depense = ?,
                        effectue = ?, emprunte = ?, est_credit = ?
                    WHERE id = ?
                '''
                values = (
                    depense.nom, depense.montant, depense.categorie, depense.date_depense,
                    depense.effectue, depense.emprunte, depense.est_credit,
                    depense.id
                )
                cursor.execute(sql, values)
                conn.commit()
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la mise à jour de la dépense: {e}")
            
    def delete_depense(self, depense_id: int):
        """Supprime une dépense"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('DELETE FROM depenses WHERE id = ?', (depense_id,))
                conn.commit()
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la suppression de la dépense: {e}")
    
    def delete_all_depenses_by_mois(self, mois_id: int):
        """Supprime toutes les dépenses d'un mois"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('DELETE FROM depenses WHERE mois_id = ?', (mois_id,))
                conn.commit()
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la suppression des dépenses: {e}")
    
    def save_config(self, key: str, value: str):
        """Sauvegarde une valeur de configuration"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    'INSERT OR REPLACE INTO config (cle, valeur) VALUES (?, ?)',
                    (key, value)
                )
                conn.commit()
        except sqlite3.Error as e:
            logger.warning(f"Erreur lors de la sauvegarde de config: {e}")
    
    def get_config(self, key: str) -> Optional[str]:
        """Récupère une valeur de configuration"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT valeur FROM config WHERE cle = ?', (key,))
                row = cursor.fetchone()
                return row[0] if row else None
        except sqlite3.Error as e:
            logger.warning(f"Erreur lors de la récupération de config: {e}")
            return None
        
    def get_mois_by_id(self, mois_id: int) -> Optional[Mois]:
        """Récupère les détails d'un mois par son ID."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT id, nom, salaire FROM mois WHERE id = ?', (mois_id,))
                row = cursor.fetchone()
                if row:
                    return Mois(id=row[0], nom=row[1], salaire=row[2])
                return None
        except sqlite3.Error as e:
            raise DatabaseError(f"Erreur lors de la récupération du mois par ID: {e}")

    def duplicate_mois(self, original_mois_id: int, new_mois_name: str) -> Result:
        """Crée une copie d'un mois existant avec toutes ses opérations."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()

                # Étape 1 : Récupérer les détails du mois original
                original_mois = self.get_mois_by_id(original_mois_id)
                if not original_mois:
                    return Result.error("Le mois original à dupliquer n'a pas été trouvé.")

                # Étape 2 : Créer le nouveau mois avec le même salaire
                # La méthode create_mois gère déjà le cas où le nom existerait
                new_mois_id = self.create_mois(new_mois_name, original_mois.salaire)

                # Étape 3 : Récupérer toutes les opérations du mois original
                original_depenses = self.get_depenses_by_mois(original_mois_id)

                # Étape 4 : Insérer des copies de ces opérations pour le nouveau mois
                for depense in original_depenses:
                    new_depense = Depense(
                        nom=depense.nom,
                        montant=depense.montant,
                        categorie=depense.categorie,
                        date_depense=depense.date_depense,
                        est_credit=depense.est_credit,
                        effectue=depense.effectue,
                        emprunte=depense.emprunte
                    )
                    self.create_depense(new_mois_id, new_depense)
                
                conn.commit()
                return Result.success(f"Mois '{original_mois.nom}' dupliqué avec succès en '{new_mois_name}'.")

        except Exception as e:
            # Lève une exception personnalisée qui sera attrapée par le modèle
            raise DatabaseError(f"Erreur lors de la duplication du mois : {e}")

# ===== SERVICE IMPORT/EXPORT =====
class ImportExportService:
    """Service pour l'import/export JSON"""
    
    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager
    
    def export_to_json(self, mois_id: int, filepath: Path) -> Result:
        """Exporte les données d'un mois (salaire et dépenses) vers un fichier JSON."""
        try:
            mois_details = self.db_manager.get_mois_by_id(mois_id)
            if not mois_details:
                return Result.error("Mois non trouvé pour l'export.")
            
            depenses = self.db_manager.get_depenses_by_mois(mois_id)
            
            # --- CORRECTION ---
            # On utilise asdict(dep) pour les dataclasses au lieu de dep._asdict()
            depenses_data = [asdict(dep) for dep in depenses]
            
            # Structure des données pour le fichier JSON
            data_to_export = {
                "mois": {
                    "nom": mois_details.nom,
                    "salaire": mois_details.salaire
                },
                "depenses": depenses_data
            }

            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data_to_export, f, indent=4, ensure_ascii=False)
            
            return Result.success(f"Export réussi vers {filepath.name}")

        except Exception as e:
            logger.error(f"Erreur inattendue lors de l'export JSON: {e}")
            return Result.error(f"Une erreur inattendue est survenue: {e}")

    def import_from_json(self, filepath: Path, new_mois_name: str) -> Result:
        """Importe un fichier JSON pour créer un nouveau mois et ses dépenses."""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if 'depenses' not in data:
                return Result.error("Structure JSON invalide : clé 'depenses' manquante.")

            salaire = data.get('mois', {}).get('salaire', 0.0)

            new_mois_id = self.db_manager.create_mois(new_mois_name, salaire)
            
            imported_count = 0
            for dep_data in data['depenses']:
                depense = Depense(
                    id=None,
                    nom=dep_data.get('nom', 'Dépense sans nom'),
                    montant=dep_data.get('montant', 0.0),
                    categorie=dep_data.get('categorie', 'Autres'),
                    effectue=dep_data.get('effectue', False),
                    emprunte=dep_data.get('emprunte', False)
                )
                self.db_manager.create_depense(new_mois_id, depense)
                imported_count += 1
            
            return Result.success(f"Import réussi: {imported_count} dépenses ajoutées à '{new_mois_name}'.")

        except FileNotFoundError:
            return Result.error("Fichier non trouvé.")
        except json.JSONDecodeError as e:
            return Result.error(f"Erreur de décodage JSON : {e}")
        except DatabaseError as e:
            return Result.error(str(e))
        except Exception as e:
            logger.error(f"Erreur inattendue lors de l'import JSON: {e}")
            return Result.error(f"Une erreur inattendue est survenue: {e}")

    # Dans utils.py, remplacez cette méthode dans la classe ImportExportService

    def import_from_excel(self, filepath: Path, new_mois_name: str, progress_callback=None) -> Result:
        """
        Lit un fichier Excel, crée un nouveau mois et y importe les opérations.
        Les crédits sont sommés et définissent le salaire initial du mois.
        """
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            sheet: Worksheet = workbook.active
            
            header_row_index = -1
            col_indices = {}

            # Recherche de la ligne d'en-tête à partir de la ligne 10
            for i in range(1, sheet.max_row + 1): # On commence à la ligne 1 pour plus de flexibilité
                row_values = [str(cell.value).strip() if cell.value else "" for cell in sheet[i]]
                if "Libellé" in row_values and "Date" in row_values:
                    header_row_index = i
                    col_indices["nom"] = row_values.index("Libellé")
                    col_indices["date"] = row_values.index("Date")
                    if "Débit euros" in row_values:
                        col_indices["debit"] = row_values.index("Débit euros")
                    if "Crédit euros" in row_values:
                        col_indices["credit"] = row_values.index("Crédit euros")
                    break
            
            if header_row_index == -1:
                return Result.error("En-tête non trouvé. Vérifiez que les colonnes 'Libellé' et 'Date' existent.")

            # On lit d'abord toutes les lignes en mémoire pour connaître le total
            rows_to_process = list(sheet.iter_rows(min_row=header_row_index + 1))
            total_rows = len(rows_to_process)
            if total_rows == 0:
                return Result.error("Aucune donnée trouvée après l'en-tête.")

            operations_a_importer: List[Depense] = []
            
            # On parcourt les lignes pré-chargées
            for row in rows_to_process: 

                cells = [cell.value for cell in row]
                nom = cells[col_indices["nom"]]
                date_val = cells[col_indices["date"]]
                debit_val = cells[col_indices.get("debit", -1)] if "debit" in col_indices else None
                credit_val = cells[col_indices.get("credit", -1)] if "credit" in col_indices else None

                if not nom or not date_val:
                    continue

                date_depense_str = ""
                if isinstance(date_val, datetime.datetime):
                    date_depense_str = date_val.strftime('%d/%m/%Y')
                else:
                    date_depense_str = str(date_val)

                montant, est_credit = 0.0, False
                try:
                    if "credit" in col_indices and credit_val and float(credit_val) > 0:
                        montant, est_credit = float(credit_val), True
                    elif "debit" in col_indices and debit_val and float(debit_val) > 0:
                        montant, est_credit = float(debit_val), False
                    else:
                        continue
                except (ValueError, TypeError):
                    logger.warning(f"Ligne {i + header_row_index + 1} ignorée: montant invalide.")
                    continue
                
                operations_a_importer.append(
                    Depense(nom=str(nom).strip(), montant=montant, date_depense=date_depense_str, est_credit=est_credit, effectue= True)
                )
            
            # 1. On calcule le total des crédits qui servira de salaire initial
            salaire_initial = sum(op.montant for op in operations_a_importer if op.est_credit)

            # 2. On crée le mois avec ce salaire
            try:
                new_mois_id = self.db_manager.create_mois(new_mois_name, salaire_initial)
            except DatabaseError as e:
                return Result.error(str(e))

            # 3. On insère toutes les opérations (crédits et débits) dans la table
            for op in operations_a_importer:
                self.db_manager.create_depense(new_mois_id, op)

            return Result.success(f"{len(operations_a_importer)} opérations importées dans '{new_mois_name}'.")

        except FileNotFoundError:
            return Result.error("Fichier Excel non trouvé.")
        except Exception as e:
            logger.error(f"Erreur inattendue lors de l'import Excel: {e}")
            return Result.error(f"Une erreur inattendue est survenue: {e}")
                        
# ===== PATTERN OBSERVER =====
class Observable:
    """Classe de base pour les objets observables"""
    
    def __init__(self):
        self._observers = []
    
    def add_observer(self, observer):
        if observer not in self._observers:
            self._observers.append(observer)
    
    def remove_observer(self, observer):
        if observer in self._observers:
            self._observers.remove(observer)
    
    def notify_observers(self, event_type: str, data: Any = None):
        for observer in self._observers:
            observer.on_model_changed(event_type, data)

class Observer:
    """Interface pour les observateurs"""
    
    def on_model_changed(self, event_type: str, data: Any):
        pass